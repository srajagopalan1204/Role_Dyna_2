
import argparse
import os
from datetime import datetime
from zoneinfo import ZoneInfo
from collections import Counter, defaultdict
import re

import pandas as pd

# =============================
# Constants & Helpers
# =============================

REQUIRED_HEADER_TOKENS = [
    "location",
    "status",
    "primary functional group",
    "secondary functional group",
]
REJECT_ROW_PREFIXES = ["Planning on Retiring", "Termed", "New Hires"]

CORE_COLS = [
    "Name",
    "Position/Title",
    "Primary Functional Group",
    "Secondary Functional Group",
    "Primary Specialization",
    "Secondary Specialization",
    "Location"
]


def tz_now_str(tz_name: str = "America/New_York"):
    now = datetime.now(ZoneInfo(tz_name))
    return now, now.strftime("%m%d%Y_%H_%M")


def load_config(path: str):
    import json
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def read_all_sheets(path: str) -> dict:
    return pd.read_excel(path, sheet_name=None, dtype=str)


def normalize_cols(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df


def detect_header_row(df: pd.DataFrame):
    """
    Scan rows to find a header line that contains the required tokens.
    Returns (header_row_index, header_values) if found, else (None, None).
    """
    for idx, row in df.iterrows():
        vals = [str(x).strip().lower() for x in row.fillna("").tolist()]
        if ("postion/title" in vals or "position/title" in vals) and all(
            any(token == v for v in vals) for token in REQUIRED_HEADER_TOKENS
        ):
            cols = [str(x).strip() for x in row.tolist()]
            return idx, cols
    return None, None


def slice_table_from_header(df: pd.DataFrame, header_idx: int, cols: list) -> pd.DataFrame:
    df2 = df.iloc[header_idx + 1 :].copy()
    df2.columns = cols
    df2 = df2.dropna(how="all")
    return normalize_cols(df2)


def coalesce_title_column(df: pd.DataFrame) -> pd.DataFrame:
    # Unify to "Position/Title"
    if "Position/Title" in df.columns and "Postion/Title" in df.columns:
        df["Position/Title"] = df["Position/Title"].fillna(df["Postion/Title"])
        df.drop(columns=["Postion/Title"], inplace=True)
    elif "Postion/Title" in df.columns and "Position/Title" not in df.columns:
        df.rename(columns={"Postion/Title": "Position/Title"}, inplace=True)
    return df


def drop_repeated_headers(df: pd.DataFrame) -> pd.DataFrame:
    """Remove rows that are header repeats inside the data region."""
    header_like = (
        (df.get("Location", "").astype(str).str.lower() == "location")
        | (df.get("Status", "").astype(str).str.lower() == "status")
        | (df.get("Primary Functional Group", "").astype(str).str.lower() == "primary functional group")
        | (df.get("Secondary Functional Group", "").astype(str).str.lower() == "secondary functional group")
        | (df.get("Position/Title", "").astype(str).str.lower().isin(["position/title","postion/title"]))
    )
    if header_like.any():
        df = df[~header_like].copy()
    return df


def filter_reject_rows(df: pd.DataFrame):
    def starts_with_banned(row) -> bool:
        for v in row:
            if pd.notna(v) and str(v).strip() != "":
                val = str(v).strip().lower()
                return any(val.startswith(b.lower()) for b in REJECT_ROW_PREFIXES)
        return False
    mask = df.apply(starts_with_banned, axis=1)
    return df[~mask].copy(), df[mask].copy()


def make_title_normalizer(tn_df: pd.DataFrame) -> dict:
    tn_df = normalize_cols(tn_df)
    cols = {c.lower(): c for c in tn_df.columns}
    raw_col = cols.get("raw") or cols.get("postion/title")
    std_col = cols.get("std") or cols.get("position/title")
    if not raw_col or not std_col:
        return {}
    return dict(
        zip(
            tn_df[raw_col].fillna("").astype(str).str.strip(),
            tn_df[std_col].fillna("").astype(str).str.strip(),
        )
    )


def map_with_errors(series: pd.Series, mapping: dict, field: str, tab: str, errors: list) -> pd.Series:
    def f(x):
        key = str(x).strip()
        if key in mapping and key != "":
            return mapping[key]
        if key == "":
            return ""
        msg = f"Error: {field} from {tab} not found -> '{key}'"
        errors.append(msg)
        return f"Error: {field} from {tab} not found"
    return series.apply(f)


def build_hndl_key(df: pd.DataFrame) -> pd.Series:
    # Key order: Position/Title | Primary Functional Group | Secondary Functional Group | Primary Specialization | Secondary Specialization
    for col in [
        "Position/Title",
        "Primary Functional Group",
        "Secondary Functional Group",
        "Primary Specialization",
        "Secondary Specialization",
    ]:
        if col not in df.columns:
            df[col] = ""
    return (
        df["Position/Title"].fillna("").astype(str).str.strip()
        + " | "
        + df["Primary Functional Group"].fillna("").astype(str).str.strip()
        + " | "
        + df["Secondary Functional Group"].fillna("").astype(str).str.strip()
        + " | "
        + df["Primary Specialization"].fillna("").astype(str).str.strip()
        + " | "
        + df["Secondary Specialization"].fillna("").astype(str).str.strip()
    )


def add_conditional_error_formatting(ws):
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import PatternFill
    from openpyxl.formatting.rule import FormulaRule

    max_row = ws.max_row
    max_col = ws.max_column
    if max_row < 2 or max_col < 1:
        return

    rng = f"A2:{get_column_letter(max_col)}{max_row}"
    # Highlight any cell whose text contains "Error"
    formula = 'ISNUMBER(SEARCH("Error",A2))'
    rule = FormulaRule(formula=[formula], stopIfTrue=False,
                       fill=PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid"))
    ws.conditional_formatting.add(rng, rule)


def norm_error(msg: str) -> str:
    """Normalize error string for grouping (strip trailing details)."""
    return re.sub(r"->.*$", "", str(msg)).strip()


def compile_error_catalog(df: pd.DataFrame) -> pd.DataFrame:
    """Scan a dataframe for 'Error:' cells and summarize counts by message."""
    errs = []
    for c in df.columns:
        s = df[c].dropna().astype(str)
        errs.extend([x for x in s if x.startswith("Error:")])
    cnt = Counter(norm_error(e) for e in errs)
    if not cnt:
        return pd.DataFrame(columns=["Error_Message", "Count"])
    return pd.DataFrame([{"Error_Message": k, "Count": v} for k, v in cnt.most_common()])


def extract_unmapped_values(errors: list) -> dict:
    """
    From the errors list (with tails '-> value'), extract unmapped lookups.
    Returns dict of lists keyed by a label.
    """
    buckets = defaultdict(set)
    for e in errors:
        m = re.search(r"Error: (.+?) from (.+?) not found -> '(.+)'", e)
        if m:
            field, tab, val = m.groups()
            key = f"Unmapped_{field}_from_{tab}"
            buckets[key].add(val)
            continue
        m2 = re.search(r"Operator ID not found -> '(.+)'", e)
        if m2:
            buckets["Unmapped_Operator_Name"].add(m2.group(1))
            continue
        m3 = re.search(r"Handler tab not found: ([A-Za-z0-9_]+)", e)
        if m3:
            buckets["Missing_Handler_Tabs"].add(m3.group(1))
            continue
        m4 = re.search(r"Role_Nm not found for key in ([A-Za-z0-9_]+): (.+)$", e)
        if m4:
            tab, key = m4.groups()
            buckets[f"Missing_Role_Key_in_{tab}"].add(key)
            continue
    return {k: sorted(v) for k, v in buckets.items()}


def drop_truly_empty_rows(df: pd.DataFrame) -> pd.DataFrame:
    """Drop rows where all CORE_COLS are empty/NaN."""
    work = df.copy()
    for c in CORE_COLS:
        if c not in work.columns:
            work[c] = ""
    mask_all_empty = work[CORE_COLS].apply(lambda r: all(str(x).strip() == "" for x in r.values), axis=1)
    return df[~mask_all_empty].copy()


# =============================
# Main
# =============================

def main():
    parser = argparse.ArgumentParser(description="Roles Dyna - Phase 1 (v1.3)")
    parser.add_argument("--config", default="config.json")
    parser.add_argument("--raw", default=None)
    parser.add_argument("--vars", default=None)
    parser.add_argument("--fgp", default="")
    parser.add_argument("--errors-only", action="store_true")
    parser.add_argument("--version", default="v_1_3")
    parser.add_argument("--outputs-dir", default=None)
    parser.add_argument("--logs-dir", default=None)
    args = parser.parse_args()

    # Sensible relative defaults (can be overridden by config or CLI)
    base_dir = os.path.dirname(os.path.abspath(__file__))
    default_raw = os.path.abspath(os.path.join(base_dir, "../../Input/Raw_input_07012025.xlsx"))
    default_vars = os.path.abspath(os.path.join(base_dir, "../../Input/Var_Trns_Mstr.xlsx"))
    default_out = os.path.abspath(os.path.join(base_dir, "../../Output"))
    default_logs = os.path.abspath(os.path.join(base_dir, "../../Logs"))

    cfg = {}
    if os.path.exists(args.config):
        try:
            cfg = load_config(args.config)
        except Exception:
            cfg = {}

    tz_name = (cfg.get("timezone") if isinstance(cfg, dict) else None) or "America/New_York"
    now, now_tag = tz_now_str(tz_name)

    raw_path = args.raw or (cfg.get("paths", {}).get("raw_input") if isinstance(cfg, dict) else None) or default_raw
    vars_path = args.vars or (cfg.get("paths", {}).get("translation_master") if isinstance(cfg, dict) else None) or default_vars
    outputs_dir = args.outputs_dir or (cfg.get("paths", {}).get("outputs_dir") if isinstance(cfg, dict) else None) or default_out
    logs_dir = args.logs_dir or (cfg.get("paths", {}).get("logs_dir") if isinstance(cfg, dict) else None) or default_logs
    fgp_filter = args.fgp or (cfg.get("options", {}).get("fgp_filter") if isinstance(cfg, dict) else "") or ""
    errors_only = bool(args.errors_only or ((cfg.get("options", {}) if isinstance(cfg, dict) else {}).get("errors_only", False)))
    version = args.version

    os.makedirs(outputs_dir, exist_ok=True)
    os.makedirs(logs_dir, exist_ok=True)

    run_log_lines = []
    def log(line: str):
        ts = now.strftime("%Y-%m-%d %H:%M:%S")
        run_log_lines.append(f"[{ts}] {line}")

    log("Run start")
    log(f"Version: {version}")
    log(f"Raw: {raw_path}")
    log(f"Vars: {vars_path}")
    log(f"FGP filter: '{fgp_filter}'")

    # --- Load translation master ---
    var_sheets = read_all_sheets(vars_path)
    Title_Normalization = var_sheets.get("Title_Normalization")
    PFGC_Conv = var_sheets.get("PFGC_Conv")
    PSP_Conv = var_sheets.get("PSP_Conv")
    SSPL_Conv = var_sheets.get("SSPL_Conv")
    Loc_Sz = var_sheets.get("Loc_Sz")
    Oper_list_saso = var_sheets.get("Oper_list_saso")
    if Oper_list_saso is None and "Oper_List_Saso" in var_sheets:
        Oper_list_saso = var_sheets["Oper_List_Saso"]
    ADDL_Keys = var_sheets.get("ADDL_Keys")

    # Maps
    pfg_map = {}
    if PFGC_Conv is not None:
        PFGC_Conv = normalize_cols(PFGC_Conv)
        cols = {c.lower(): c for c in PFGC_Conv.columns}
        long_col = cols.get("orig_primary_functional_group") or cols.get("primary functional group") or list(PFGC_Conv.columns)[0]
        short_col = cols.get("shrt_func") or list(PFGC_Conv.columns)[1]
        pfg_map = dict(zip(PFGC_Conv[long_col].fillna("").astype(str).str.strip(),
                           PFGC_Conv[short_col].fillna("").astype(str).str.strip()))

    psp_map = {}
    if PSP_Conv is not None:
        PSP_Conv = normalize_cols(PSP_Conv)
        cols = {c.lower(): c for c in PSP_Conv.columns}
        long_col = cols.get("primary specialization") or list(PSP_Conv.columns)[0]
        short_col = cols.get("shrtpspec") or list(PSP_Conv.columns)[1]
        psp_map = dict(zip(PSP_Conv[long_col].fillna("").astype(str).str.strip(),
                           PSP_Conv[short_col].fillna("").astype(str).str.strip()))

    sspl_map = {}
    if SSPL_Conv is not None:
        SSPL_Conv = normalize_cols(SSPL_Conv)
        cols = {c.lower(): c for c in SSPL_Conv.columns}
        long_col = cols.get("secondary specialization") or list(SSPL_Conv.columns)[0]
        short_col = cols.get("shrt_sspl") or list(SSPL_Conv.columns)[1]
        sspl_map = dict(zip(SSPL_Conv[long_col].fillna("").astype(str).str.strip(),
                            SSPL_Conv[short_col].fillna("").astype(str).str.strip()))

    title_map = make_title_normalizer(Title_Normalization) if Title_Normalization is not None else {}

    # Operator map
    oper_map = {}
    if Oper_list_saso is not None:
        df_op = normalize_cols(Oper_list_saso)
        cols = {c.lower(): c for c in df_op.columns}
        name_col = cols.get("name") or list(df_op.columns)[0]
        oper_col = cols.get("operator") or list(df_op.columns)[0]
        if name_col and oper_col:
            oper_map = dict(
                zip(
                    df_op[name_col].fillna("").astype(str).str.strip(),
                    df_op[oper_col].fillna("").astype(str).str.strip(),
                )
            )

    # Location size map
    loc_size_map = {}
    if Loc_Sz is not None:
        df_ls = normalize_cols(Loc_Sz)
        cols = {c.lower(): c for c in df_ls.columns}
        loc_col = cols.get("location") or list(df_ls.columns)[0]
        size_col = cols.get("size") or list(df_ls.columns)[1]
        loc_size_map = dict(
            zip(
                df_ls[loc_col].fillna("").astype(str).str.strip(),
                df_ls[size_col].fillna("").astype(str).str.strip(),
            )
        )

    # ADDL Keys
    addl_keys = []
    if ADDL_Keys is not None:
        df_k = normalize_cols(ADDL_Keys)
        if "Key_name" in df_k.columns:
            addl_keys = [k for k in df_k["Key_name"].dropna().astype(str).str.strip().tolist() if k]

    # --- RAW input: stitch tables from all sheets ---
    raw_book = read_all_sheets(raw_path)
    tables = []
    rejected_all = []
    for sh_name, df in raw_book.items():
        df = df.fillna("")
        header_idx, cols = detect_header_row(df)
        if header_idx is None:
            log(f"[{sh_name}] No header detected; skipping sheet.")
            continue
        tbl = slice_table_from_header(df, header_idx, cols)
        tbl = coalesce_title_column(tbl)
        tbl = drop_repeated_headers(tbl)
        tbl, rejected = filter_reject_rows(tbl)
        tbl["__source_sheet"] = sh_name
        rejected["__source_sheet"] = sh_name
        tables.append(tbl)
        if not rejected.empty:
            rejected_all.append(rejected)

    if not tables:
        raise SystemExit("No valid tables detected in RAW input.")

    raw_df = pd.concat(tables, ignore_index=True)
    raw_df = drop_truly_empty_rows(raw_df)

    # Optional filter by FGP (accept long or short form)
    if fgp_filter:
        fgp_lower = fgp_filter.strip().lower()
        raw_df = raw_df[
            raw_df.get("Primary Functional Group","").astype(str).str.lower().eq(fgp_lower)
            | raw_df.get("Primary Functional Group","").map(lambda x: pfg_map.get(str(x).strip(), "")).str.lower().eq(fgp_lower)
        ]

    # Title normalization
    errors = []
    if title_map and "Position/Title" in raw_df.columns:
        raw_df["Position/Title"] = raw_df["Position/Title"].apply(lambda x: title_map.get(str(x).strip(), str(x).strip()))

    # Short forms
    if "Primary Functional Group" in raw_df.columns and pfg_map:
        raw_df["Shrt_Func"] = map_with_errors(
            raw_df["Primary Functional Group"], pfg_map, "Primary Functional Group", "PFGC_Conv", errors
        )
    else:
        raw_df["Shrt_Func"] = ""

    if "Primary Specialization" in raw_df.columns and psp_map:
        raw_df["ShrtPSpec"] = map_with_errors(
            raw_df["Primary Specialization"], psp_map, "Primary Specialization", "PSP_Conv", errors
        )
    else:
        raw_df["ShrtPSpec"] = ""

    if "Secondary Specialization" in raw_df.columns and sspl_map:
        raw_df["Shrt_SSpl"] = map_with_errors(
            raw_df["Secondary Specialization"], sspl_map, "Secondary Specialization", "SSPL_Conv", errors
        )
    else:
        raw_df["Shrt_SSpl"] = ""

    # Location size
    if loc_size_map and "Location" in raw_df.columns:
        raw_df["Location_Size"] = raw_df["Location"].apply(lambda x: loc_size_map.get(str(x).strip(), ""))

    # Operator id
    if oper_map and "Name" in raw_df.columns:
        def map_oper(nm):
            key = str(nm).strip()
            if key in oper_map and key != "":
                return oper_map[key]
            if key == "":
                return ""
            errors.append(f"Operator ID not found -> '{key}'")
            return "Error: Operator ID not found"
        raw_df["oper"] = raw_df["Name"].apply(map_oper)

    # Build handler key
    raw_df["__hndl_key"] = build_hndl_key(raw_df)

    # =============================
    # Vectorized Role Lookup (cached handler tabs)
    # =============================
    # Determine pf_short per row
    raw_df["__pf_short"] = raw_df.apply(
        lambda r: pfg_map.get(str(r.get("Primary Functional Group","")).strip(), str(r.get("Shrt_Func","")).strip()), axis=1
    )

    # Cache handler dfs per pf_short, standardize ROLE_Nm col, build keys once
    handlers_all = []
    missing_tabs = set()
    for pf_short in sorted(set([x for x in raw_df["__pf_short"].astype(str).tolist() if x])):
        tab_name = f"{pf_short}_Hndl"
        hndl_df = var_sheets.get(tab_name)
        if hndl_df is None:
            missing_tabs.add(tab_name)
            continue
        hndl_df = normalize_cols(hndl_df)
        hndl_df = coalesce_title_column(hndl_df)
        hndl_df["__hndl_key"] = build_hndl_key(hndl_df)

        # Find role column
        role_col_name = None
        for c in hndl_df.columns:
            lc = c.strip().lower()
            if lc in ("role_nm","role name","role"):
                role_col_name = c
                break
        if role_col_name is None:
            # Still create placeholder to trigger "column missing" error later
            hndl_df["ROLE_Nm"] = ""
        else:
            hndl_df.rename(columns={role_col_name: "ROLE_Nm"}, inplace=True)

        handlers_all.append(hndl_df[["__hndl_key", "ROLE_Nm"]])

    if handlers_all:
        handlers_all_df = pd.concat(handlers_all, ignore_index=True)
    else:
        handlers_all_df = pd.DataFrame(columns=["__hndl_key","ROLE_Nm"])

    # Merge once
    merged = raw_df.merge(handlers_all_df, on="__hndl_key", how="left", suffixes=("", "_hndl"))

    # Fill ROLE_Nm with appropriate errors
    def role_error_for_row(row):
        pf_short = str(row.get("__pf_short","")).strip()
        if not pf_short:
            return "Error: Primary Functional Group short code missing"
        tab_name = f"{pf_short}_Hndl"
        if tab_name in missing_tabs:
            return f"Error: Handler tab not found: {tab_name}"
        if pd.isna(row.get("ROLE_Nm","")) or str(row.get("ROLE_Nm","")).strip() == "":
            return f"Error: Role_Nm not found in {tab_name}"
        return str(row["ROLE_Nm"]).strip()

    merged["ROLE_Nm"] = merged.apply(role_error_for_row, axis=1)

    # ADDL Keys placeholders (create columns but leave blank in Phase 1)
    for k in addl_keys:
        if k not in merged.columns:
            merged[k] = ""

    # Final master
    master_df = merged.drop(columns=["__pf_short"], errors="ignore").copy()
    master_df = drop_truly_empty_rows(master_df)

    # Exceptions detection
    excp_mask = master_df.apply(lambda r: any(isinstance(v, str) and v.startswith("Error:") for v in r.values), axis=1)
    excp_rows = master_df[excp_mask].copy()

    if errors_only:
        master_df = master_df[excp_mask].copy()

    # Basic Summary (starter)
    def safe_count(df, cols):
        existing = [c for c in cols if c in df.columns]
        if not existing:
            return pd.DataFrame()
        return df.groupby(existing, dropna=False).size().reset_index(name="Count")

    summary_blocks = []
    s1 = safe_count(master_df, ["Primary Functional Group", "ROLE_Nm"])
    if not s1.empty:
        s1.insert(0, "__metric", "Count by PFG + ROLE_Nm")
        summary_blocks.append(s1)
    s2 = safe_count(master_df, ["Primary Functional Group", "Position/Title"])
    if not s2.empty:
        s2.insert(0, "__metric", "Count by PFG + Position/Title")
        summary_blocks.append(s2)
    summary_df = pd.concat(summary_blocks, ignore_index=True) if summary_blocks else pd.DataFrame({"__metric": []})

    # Run_Rep
    run_rep = pd.DataFrame(
        [{
            "Run_Timestamp": now.strftime("%m%d%Y_%H_%M"),
            "Timezone": tz_name,
            "Version": version,
            "Raw_File": raw_path,
            "Var_Master_File": vars_path,
            "FGP_Filter": fgp_filter,
            "Total_Records": len(master_df),
            "Exception_Rows": len(excp_rows),
        }]
    )

    # Error Catalog + Unmapped tabs
    error_catalog = compile_error_catalog(master_df)

    # Collect all errors from mapping steps for unmapped reporting
    mapping_errors = [e for e in globals().get("errors", [])] if "errors" in globals() else []
    # Also scan master_df cells for 'Error:' tails to extract values
    for c in master_df.columns:
        s = master_df[c].dropna().astype(str)
        mapping_errors.extend([x for x in s if x.startswith("Error:")])

    unmapped_dict = extract_unmapped_values(mapping_errors)
    unmapped_tabs = {k: pd.DataFrame({k: v}) for k, v in unmapped_dict.items()}

    # Write Excel with conditional formatting (fast)
    out_name = f"Roles_Output_{version}_{now.strftime('%d%m%y_%H_%M')}.xlsx"
    out_path = os.path.join(outputs_dir, out_name)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        master_df.to_excel(writer, index=False, sheet_name="Master_Roles_by_User")
        summary_df.to_excel(writer, index=False, sheet_name="Summary")
        run_rep.to_excel(writer, index=False, sheet_name="Run_Rep")
        excp_rows.to_excel(writer, index=False, sheet_name="Excp_Rep")
        error_catalog.to_excel(writer, index=False, sheet_name="Error_Catalog")
        for tab, df_tab in unmapped_tabs.items():
            sheet_name = tab[:31]  # Excel limit
            df_tab.to_excel(writer, index=False, sheet_name=sheet_name)

        wb = writer.book
        ws = writer.sheets["Master_Roles_by_User"]
        add_conditional_error_formatting(ws)

    # Append to external log (append at bottom for Phase 1)
    log_path = os.path.join(logs_dir, "Roles_Dyn_cons_Log.txt")
    try:
        mode = "a" if os.path.exists(log_path) else "w"
        with open(log_path, mode, encoding="utf-8") as lf:
            lf.write("\n".join(run_log_lines) + "\n")
    except Exception as e:
        print(f"WARNING: could not write log file: {e}")

    print(f"Output written: {out_path}")
    print(f"Exceptions: {len(excp_rows)}")


if __name__ == "__main__":
    main()
