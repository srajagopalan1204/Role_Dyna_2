
import argparse
import os
import re
from datetime import datetime
from zoneinfo import ZoneInfo
import pandas as pd

# ---- Constants ----
REQUIRED_HEADER_TOKENS = ["location", "status", "primary functional group", "secondary functional group"]
REJECT_ROW_PREFIXES = ["Planning on Retiring", "Termed", "New Hires"]

# ---- Time helpers ----
def tz_now_str(tz_name: str = "America/New_York"):
    now = datetime.now(ZoneInfo(tz_name))
    return now, now.strftime("%m%d%Y_%H_%M")

# ---- Config ----
def load_config(path: str):
    import json
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

# ---- I/O helpers ----
def read_all_sheets(path: str) -> dict:
    return pd.read_excel(path, sheet_name=None, dtype=str)

def normalize_cols(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df

def detect_header_row(df: pd.DataFrame):
    for idx, row in df.iterrows():
        vals = [str(x).strip().lower() for x in row.fillna("").tolist()]
        if ("postion/title" in vals or "position/title" in vals) and all(any(tok == v for v in vals) for tok in REQUIRED_HEADER_TOKENS):
            cols = [str(x).strip() for x in row.tolist()]
            return idx, cols
    return None, None

def slice_table_from_header(df: pd.DataFrame, header_idx: int, cols: list) -> pd.DataFrame:
    df2 = df.iloc[header_idx+1:].copy()
    df2.columns = cols
    df2 = df2.dropna(how="all")
    return normalize_cols(df2)

def coalesce_title_column(df: pd.DataFrame) -> pd.DataFrame:
    if "Position/Title" in df.columns and "Postion/Title" in df.columns:
        df["Position/Title"] = df["Position/Title"].fillna(df["Postion/Title"])
        df.drop(columns=["Postion/Title"], inplace=True)
    elif "Postion/Title" in df.columns and "Position/Title" not in df.columns:
        df.rename(columns={"Postion/Title": "Position/Title"}, inplace=True)
    return df

def filter_reject_rows(df: pd.DataFrame):
    def starts_with_banned(row) -> bool:
        for v in row:
            if pd.notna(v) and str(v).strip():
                val = str(v).strip().lower()
                return any(val.startswith(b.lower()) for b in REJECT_ROW_PREFIXES)
        return False
    mask = df.apply(starts_with_banned, axis=1)
    return df[~mask].copy(), df[mask].copy()

def make_title_normalizer(tn_df: pd.DataFrame) -> dict:
    tn_df = normalize_cols(tn_df)
    cols = {c.lower(): c for c in tn_df.columns}
    raw_col = cols.get("raw") or cols.get("postion/title")
    std_col = cols.get("std") or cols.get("position/title")
    if not raw_col or not std_col:
        return {}
    return dict(zip(tn_df[raw_col].fillna("").astype(str).str.strip(),
                    tn_df[std_col].fillna("").astype(str).str.strip()))

def map_with_errors(series: pd.Series, mapping: dict, field: str, tab: str, errors: list) -> pd.Series:
    def f(x):
        key = str(x).strip()
        if key in mapping and key != "":
            return mapping[key]
        if key == "":
            return ""
        errors.append(f"Error: {field} from {tab} not found -> '{key}'")
        return f"Error: {field} from {tab} not found"
    return series.apply(f)

def build_hndl_key(df: pd.DataFrame) -> pd.Series:
    for col in ["Position/Title","Primary Functional Group","Secondary Functional Group","Primary Specialization","Secondary Specialization"]:
        if col not in df.columns:
            df[col] = ""
    return (df["Position/Title"].fillna("").astype(str).str.strip() + " | " +
            df["Primary Functional Group"].fillna("").astype(str).str.strip() + " | " +
            df["Secondary Functional Group"].fillna("").astype(str).str.strip() + " | " +
            df["Primary Specialization"].fillna("").astype(str).str.strip() + " | " +
            df["Secondary Specialization"].fillna("").astype(str).str.strip())

# ---- Operator matching helpers ----
def _clean_name(s: str, case_insensitive: bool = True) -> str:
    s = str(s or "").strip()
    s = s.lower() if case_insensitive else s
    s = re.sub(r"[^\w\s-]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def _strip_prefixes(tokens, prefixes):
    if tokens and tokens[0] in prefixes:
        return tokens[1:]
    return tokens

def make_name_keys(raw: str, opts: dict) -> list:
    s = _clean_name(raw, opts.get("case_insensitive", True))
    tokens = s.split()
    prefixes = [p.lower() for p in opts.get("strip_prefixes", [])]
    tokens = _strip_prefixes(tokens, prefixes)
    if not tokens:
        return [""]
    keys = []
    keys.append(" ".join(tokens))                 # full
    if len(tokens) > 1:
        keys.append(tokens[0] + " " + tokens[-1]) # first last
        keys.append(tokens[-1] + " " + tokens[0]) # last first
        keys.append(tokens[0][0] + " " + tokens[-1]) # initial last
    if len(tokens) > 1 and "-" in tokens[-1]:
        keys.append(tokens[0] + " " + tokens[-1].replace("-", " "))
    return list(dict.fromkeys(keys))

def build_operator_lookup(df_op: pd.DataFrame, opts: dict):
    df_op = normalize_cols(df_op)
    cols = {c.lower(): c for c in df_op.columns}
    name_col = cols.get("name") or list(df_op.columns)[0]
    oper_col = cols.get("operator") or cols.get("role name") or cols.get("profile") or list(df_op.columns)[0]
    lookup = {}
    for _, row in df_op.iterrows():
        name = str(row.get(name_col, "")).strip()
        oper = str(row.get(oper_col, "")).strip()
        if not name:
            continue
        for k in make_name_keys(name, opts):
            if k and k not in lookup:
                lookup[k] = oper
    return lookup

def find_operator(name_raw: str, lookup: dict, opts: dict):
    keys = make_name_keys(name_raw, opts)
    for k in keys:
        if k in lookup:
            return lookup[k], "exact:" + k
    if opts.get("drop_middle_names", True) and len(keys) >= 2:
        k = keys[1]
        if k in lookup:
            return lookup[k], "first_last"
    if opts.get("use_fuzzy", True):
        try:
            from difflib import SequenceMatcher
            best_k, best_ratio = None, 0.0
            for k in lookup.keys():
                r = SequenceMatcher(None, k, keys[0]).ratio()
                if r > best_ratio:
                    best_ratio, best_k = r, k
            if best_ratio >= float(opts.get("fuzzy_threshold", 0.88)):
                return lookup[best_k], f"fuzzy:{best_ratio:.2f}<-{best_k}"
        except Exception:
            pass
    return None, ""

# ---- Main ----
def main():
    parser = argparse.ArgumentParser(description="Roles Dyna - Phase 1 (v1.1)")
    parser.add_argument("--config", default="config.json")
    parser.add_argument("--raw", default=None)
    parser.add_argument("--vars", default=None)
    parser.add_argument("--fgp", default="")
    parser.add_argument("--errors-only", action="store_true")
    parser.add_argument("--version", default="v_1_1")
    parser.add_argument("--outputs-dir", default=None)
    parser.add_argument("--logs-dir", default=None)
    args = parser.parse_args()

    cfg = load_config(args.config) if os.path.exists(args.config) else {}
    tz_name = cfg.get("timezone") or "America/New_York"
    now, now_tag = tz_now_str(tz_name)

    raw_path = args.raw or cfg.get("paths", {}).get("raw_input")
    vars_path = args.vars or cfg.get("paths", {}).get("translation_master")
    outputs_dir = args.outputs_dir or cfg.get("paths", {}).get("outputs_dir", ".")
    logs_dir = args.logs_dir or cfg.get("paths", {}).get("logs_dir", ".")
    fgp_filter = args.fgp or cfg.get("options", {}).get("fgp_filter", "")
    errors_only = args.errors_only or cfg.get("options", {}).get("errors_only", False)
    version = args.version
    oper_opts = cfg.get("operator_match", {})

    os.makedirs(outputs_dir, exist_ok=True)
    os.makedirs(logs_dir, exist_ok=True)

    run_log_lines = []
    def log(line: str):
        ts = now.strftime("%Y-%m-%d %H:%M:%S")
        run_log_lines.append(f"[{ts}] {line}")

    log("Run start")
    log(f"Version: {version}")
    log(f"Raw: {raw_path}")
    log(f"Vars: {vars_path}")
    log(f"FGP filter: '{fgp_filter}'")

    # --- Load translation master ---
    var_sheets = read_all_sheets(vars_path)
    Title_Normalization = var_sheets.get("Title_Normalization")
    PFGC_Conv = var_sheets.get("PFGC_Conv")
    PSP_Conv = var_sheets.get("PSP_Conv")
    SSPL_Conv = var_sheets.get("SSPL_Conv")
    Loc_Sz = var_sheets.get("Loc_Sz")
    # FIX: no boolean "or" on DataFrames
    Oper_list_saso = var_sheets.get("Oper_list_saso")
    if Oper_list_saso is None and "Oper_List_Saso" in var_sheets:
        Oper_list_saso = var_sheets["Oper_List_Saso"]
    ADDL_Keys = var_sheets.get("ADDL_Keys")

    # Maps
    pfg_map = {}
    if PFGC_Conv is not None:
        PFGC_Conv = normalize_cols(PFGC_Conv)
        cols = {c.lower(): c for c in PFGC_Conv.columns}
        long_col = cols.get("orig_primary_functional_group") or cols.get("primary functional group") or list(PFGC_Conv.columns)[0]
        short_col = cols.get("shrt_func") or list(PFGC_Conv.columns)[1]
        pfg_map = dict(zip(PFGC_Conv[long_col].fillna("").astype(str).str.strip(),
                           PFGC_Conv[short_col].fillna("").astype(str).str.strip()))

    psp_map = {}
    if PSP_Conv is not None:
        PSP_Conv = normalize_cols(PSP_Conv)
        cols = {c.lower(): c for c in PSP_Conv.columns}
        long_col = cols.get("primary specialization") or list(PSP_Conv.columns)[0]
        short_col = cols.get("shrtpspec") or list(PSP_Conv.columns)[1]
        psp_map = dict(zip(PSP_Conv[long_col].fillna("").astype(str).str.strip(),
                           PSP_Conv[short_col].fillna("").astype(str).str.strip()))

    sspl_map = {}
    if SSPL_Conv is not None:
        SSPL_Conv = normalize_cols(SSPL_Conv)
        cols = {c.lower(): c for c in SSPL_Conv.columns}
        long_col = cols.get("secondary specialization") or list(SSPL_Conv.columns)[0]
        short_col = cols.get("shrt_sspl") or list(SSPL_Conv.columns)[1]
        sspl_map = dict(zip(SSPL_Conv[long_col].fillna("").astype(str).str.strip(),
                            SSPL_Conv[short_col].fillna("").astype(str).str.strip()))

    title_map = make_title_normalizer(Title_Normalization) if Title_Normalization is not None else {}

    # Operator lookup
    oper_lookup = {}
    if Oper_list_saso is not None:
        oper_lookup = build_operator_lookup(Oper_list_saso, oper_opts)

    # Location size map
    loc_size_map = {}
    if Loc_Sz is not None:
        df_ls = normalize_cols(Loc_Sz)
        cols = {c.lower(): c for c in df_ls.columns}
        loc_col = cols.get("location") or list(df_ls.columns)[0]
        size_col = cols.get("size") or list(df_ls.columns)[1]
        loc_size_map = dict(zip(df_ls[loc_col].fillna("").astype(str).str.strip(),
                                df_ls[size_col].fillna("").astype(str).str.strip()))

    # ADDL Keys
    addl_keys = []
    if ADDL_Keys is not None:
        df_k = normalize_cols(ADDL_Keys)
        if "Key_name" in df_k.columns:
            addl_keys = [k for k in df_k["Key_name"].dropna().astype(str).str.strip().tolist() if k]

    # --- RAW input: stitch tables from all sheets ---
    raw_book = read_all_sheets(raw_path)
    tables = []
    rejected_all = []
    for sh_name, df in raw_book.items():
        df = df.fillna("")
        header_idx, cols = detect_header_row(df)
        if header_idx is None:
            log(f"[{sh_name}] No header detected; skipping sheet.")
            continue
        tbl = slice_table_from_header(df, header_idx, cols)
        tbl = coalesce_title_column(tbl)
        tbl, rejected = filter_reject_rows(tbl)
        tbl["__source_sheet"] = sh_name
        rejected["__source_sheet"] = sh_name
        tables.append(tbl)
        if not rejected.empty:
            rejected_all.append(rejected)

    if not tables:
        raise SystemExit("No valid tables detected in RAW input.")

    raw_df = pd.concat(tables, ignore_index=True)
    rejected_df = pd.concat(rejected_all, ignore_index=True) if rejected_all else pd.DataFrame(columns=["__source_sheet"])

    # Optional FGP filter
    if fgp_filter:
        fgp_lower = fgp_filter.strip().lower()
        raw_df = raw_df[
            raw_df["Primary Functional Group"].astype(str).str.lower().eq(fgp_lower) |
            raw_df["Primary Functional Group"].map(lambda x: pfg_map.get(str(x).strip(), "")).str.lower().eq(fgp_lower)
        ]

    # Title normalization
    errors = []
    if title_map and "Position/Title" in raw_df.columns:
        raw_df["Position/Title"] = raw_df["Position/Title"].apply(lambda x: title_map.get(str(x).strip(), str(x).strip()))

    # Short forms
    if "Primary Functional Group" in raw_df.columns and pfg_map:
        raw_df["Shrt_Func"] = map_with_errors(raw_df["Primary Functional Group"], pfg_map, "Primary Functional Group", "PFGC_Conv", errors)
    else:
        raw_df["Shrt_Func"] = ""
    if "Primary Specialization" in raw_df.columns and psp_map:
        raw_df["ShrtPSpec"] = map_with_errors(raw_df["Primary Specialization"], psp_map, "Primary Specialization", "PSP_Conv", errors)
    else:
        raw_df["ShrtPSpec"] = ""
    if "Secondary Specialization" in raw_df.columns and sspl_map:
        raw_df["Shrt_SSpl"] = map_with_errors(raw_df["Secondary Specialization"], sspl_map, "Secondary Specialization", "SSPL_Conv", errors)
    else:
        raw_df["Shrt_SSpl"] = ""

    # Location size
    if loc_size_map and "Location" in raw_df.columns:
        raw_df["Location_Size"] = raw_df["Location"].apply(lambda x: loc_size_map.get(str(x).strip(), ""))

    # Operator id with robust matching
    raw_df["oper"] = ""
    raw_df["oper_match_method"] = ""
    if oper_lookup:
        for i, nm in raw_df.get("Name", pd.Series([""]*len(raw_df))).astype(str).items():
            oper_val, how = find_operator(nm, oper_lookup, oper_opts)
            if oper_val:
                raw_df.at[i, "oper"] = oper_val
                raw_df.at[i, "oper_match_method"] = how
            else:
                errors.append(f"Operator ID not found -> '{nm}'")
                raw_df.at[i, "oper"] = "Error: Operator ID not found"
                raw_df.at[i, "oper_match_method"] = ""

    # Build handler key
    raw_df["__hndl_key"] = build_hndl_key(raw_df)

    # Role lookup via *_Hndl based on short func
    role_col = []
    for _, row in raw_df.iterrows():
        pf_long = str(row.get("Primary Functional Group","")).strip()
        pf_short = pfg_map.get(pf_long, str(row.get("Shrt_Func","")).strip())
        if not pf_short:
            role_col.append("Error: Primary Functional Group short code missing")
            continue
        tab_name = f"{pf_short}_Hndl"
        hndl_df = var_sheets.get(tab_name)
        if hndl_df is None:
            errors.append(f"Handler tab not found: {tab_name}")
            role_col.append(f"Error: Handler tab not found: {tab_name}")
            continue
        hndl_df = normalize_cols(hndl_df)
        hndl_df = coalesce_title_column(hndl_df)
        hndl_df["__hndl_key"] = build_hndl_key(hndl_df)
        key = row["__hndl_key"]
        sub = hndl_df[hndl_df["__hndl_key"] == key]
        if sub.empty:
            errors.append(f"Role_Nm not found for key in {tab_name}: {key}")
            role_col.append(f"Error: Role_Nm not found in {tab_name}")
        else:
            role_col_name = None
            for c in sub.columns:
                if c.strip().lower() in ("role_nm","role name","role"):
                    role_col_name = c
                    break
            if not role_col_name:
                errors.append(f"Role column not found in {tab_name}")
                role_col.append(f"Error: Role_Nm column missing in {tab_name}")
            else:
                role_col.append(str(sub.iloc[0][role_col_name]).strip())
    raw_df["ROLE_Nm"] = role_col

    # ADDL Keys placeholders
    for k in addl_keys:
        if k not in raw_df.columns:
            raw_df[k] = ""

    # Build Master and Exceptions
    master_df = raw_df.copy()
    excp_mask = master_df.apply(lambda r: any(isinstance(v, str) and "Error:" in v for v in r.values), axis=1)
    excp_rows = master_df[excp_mask].copy()
    if errors_only:
        master_df = master_df[excp_mask].copy()

    # Basic Summary
    def safe_count(df, cols):
        existing = [c for c in cols if c in df.columns]
        if not existing: return pd.DataFrame()
        return df.groupby(existing, dropna=False).size().reset_index(name="Count")

    summary_blocks = []
    s1 = safe_count(master_df, ["Primary Functional Group", "ROLE_Nm"])
    if not s1.empty:
        s1.insert(0, "__metric", "Count by PFG + ROLE_Nm")
        summary_blocks.append(s1)
    s2 = safe_count(master_df, ["Primary Functional Group","Position/Title"])
    if not s2.empty:
        s2.insert(0, "__metric", "Count by PFG + Position/Title")
        summary_blocks.append(s2)
    summary_df = pd.concat(summary_blocks, ignore_index=True) if summary_blocks else pd.DataFrame({"__metric":[]})

    # Run_Rep
    run_rep = pd.DataFrame([{
        "Run_Timestamp": now.strftime("%m%d%Y_%H_%M"),
        "Timezone": tz_name,
        "Version": version,
        "Raw_File": raw_path,
        "Var_Master_File": vars_path,
        "FGP_Filter": fgp_filter,
        "Total_Records": len(raw_df),
        "Master_Records_Written": len(master_df),
        "Exception_Rows": len(excp_rows),
        "Operator_Matches": int((raw_df["oper"] != "").sum() - (raw_df["oper"].astype(str).str.startswith("Error:").sum()))
    }])

    # Write Excel with error highlighting
    out_name = f"Roles_Output_{version}_{now.strftime('%d%m%y_%H_%M')}.xlsx"
    out_path = os.path.join(outputs_dir, out_name)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        master_df.to_excel(writer, index=False, sheet_name="Master_Roles_by_User")
        summary_df.to_excel(writer, index=False, sheet_name="Summary")
        run_rep.to_excel(writer, index=False, sheet_name="Run_Rep")
        excp_rows.to_excel(writer, index=False, sheet_name="Excp_Rep")
        wb = writer.book
        from openpyxl.styles import PatternFill
        ws = wb["Master_Roles_by_User"]
        fill = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid")
        for row in ws.iter_rows():
            for cell in row:
                try:
                    if isinstance(cell.value, str) and "Error" in cell.value:
                        cell.fill = fill
                except Exception:
                    pass

    # External log
    log_path = os.path.join(logs_dir, "Roles_Dyn_cons_Log.txt")
    try:
        mode = "a" if os.path.exists(log_path) else "w"
        with open(log_path, mode, encoding="utf-8") as lf:
            lf.write("\n".join(run_log_lines) + "\n")
    except Exception as e:
        print(f"WARNING: could not write log file: {e}")

    print(f"Output written: {out_path}")
    print(f"Exceptions: {len(excp_rows)}")

if __name__ == "__main__":
    main()
