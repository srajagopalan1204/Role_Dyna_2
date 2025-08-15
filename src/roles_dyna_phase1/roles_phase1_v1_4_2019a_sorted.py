
import argparse
import os
from datetime import datetime
from zoneinfo import ZoneInfo
from collections import Counter, defaultdict
import re

import pandas as pd

# =============================
# Constants & Helpers
# =============================

REQUIRED_HEADER_TOKENS = [
    "location",
    "status",
    "primary functional group",
    "secondary functional group",
]
REJECT_ROW_PREFIXES = ["Planning on Retiring", "Termed", "New Hires"]

CORE_COLS = [
    "Name",
    "Position/Title",
    "Primary Functional Group",
    "Secondary Functional Group",
    "Primary Specialization",
    "Secondary Specialization",
    "Location"
]


def tz_now_str(tz_name: str = "America/New_York"):
    now = datetime.now(ZoneInfo(tz_name))
    return now, now.strftime("%m%d%Y_%H_%M")


def load_config(path: str):
    import json
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def read_all_sheets(path: str) -> dict:
    return pd.read_excel(path, sheet_name=None, dtype=str)


def normalize_cols(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df


def detect_header_row(df: pd.DataFrame):
    """
    Scan rows to find a header line that contains the required tokens.
    Returns (header_row_index, header_values) if found, else (None, None).
    """
    for idx, row in df.iterrows():
        vals = [str(x).strip().lower() for x in row.fillna("").tolist()]
        if ("postion/title" in vals or "position/title" in vals) and all(
            any(token == v for v in vals) for token in REQUIRED_HEADER_TOKENS
        ):
            cols = [str(x).strip() for x in row.tolist()]
            return idx, cols
    return None, None


def slice_table_from_header(df: pd.DataFrame, header_idx: int, cols: list) -> pd.DataFrame:
    df2 = df.iloc[header_idx + 1 :].copy()
    df2.columns = cols
    df2 = df2.dropna(how="all")
    return normalize_cols(df2)


def coalesce_title_column(df: pd.DataFrame) -> pd.DataFrame:
    # Unify to "Position/Title"
    if "Position/Title" in df.columns and "Postion/Title" in df.columns:
        df["Position/Title"] = df["Position/Title"].fillna(df["Postion/Title"])
        df.drop(columns=["Postion/Title"], inplace=True)
    elif "Postion/Title" in df.columns and "Position/Title" not in df.columns:
        df.rename(columns={"Postion/Title": "Position/Title"}, inplace=True)
    return df


def drop_repeated_headers(df: pd.DataFrame) -> pd.DataFrame:
    """Remove rows that are header repeats inside the data region."""
    header_like = (
        (df.get("Location", "").astype(str).str.lower() == "location")
        | (df.get("Status", "").astype(str).str.lower() == "status")
        | (df.get("Primary Functional Group", "").astype(str).str.lower() == "primary functional group")
        | (df.get("Secondary Functional Group", "").astype(str).str.lower() == "secondary functional group")
        | (df.get("Position/Title", "").astype(str).str.lower().isin(["position/title","postion/title"]))
    )
    if header_like.any():
        df = df[~header_like].copy()
    return df


def filter_reject_rows(df: pd.DataFrame):
    def starts_with_banned(row) -> bool:
        for v in row:
            if pd.notna(v) and str(v).strip() != "":
                val = str(v).strip().lower()
                return any(val.startswith(b.lower()) for b in REJECT_ROW_PREFIXES)
        return False
    mask = df.apply(starts_with_banned, axis=1)
    return df[~mask].copy(), df[mask].copy()


def make_title_normalizer(tn_df: pd.DataFrame) -> dict:
    tn_df = normalize_cols(tn_df)
    cols = {c.lower(): c for c in tn_df.columns}
    raw_col = cols.get("raw") or cols.get("postion/title")
    std_col = cols.get("std") or cols.get("position/title")
    if not raw_col or not std_col:
        return {}
    return dict(
        zip(
            tn_df[raw_col].fillna("").astype(str).str.strip(),
            tn_df[std_col].fillna("").astype(str).str.strip(),
        )
    )


def map_with_errors(series: pd.Series, mapping: dict, field: str, tab: str, errors: list) -> pd.Series:
    def f(x):
        key = str(x).strip()
        if key in mapping and key != "":
            return mapping[key]
        if key == "":
            return ""
        msg = f"Error: {field} from {tab} not found -> '{key}'"
        errors.append(msg)
        return f"Error: {field} from {tab} not found"
    return series.apply(f)


def build_hndl_key(df: pd.DataFrame) -> pd.Series:
    # Key order: Position/Title | Primary Functional Group | Secondary Functional Group | Primary Specialization | Secondary Specialization
    for col in [
        "Position/Title",
        "Primary Functional Group",
        "Secondary Functional Group",
        "Primary Specialization",
        "Secondary Specialization",
    ]:
        if col not in df.columns:
            df[col] = ""
    return (
        df["Position/Title"].fillna("").astype(str).str.strip()
        + " | "
        + df["Primary Functional Group"].fillna("").astype(str).str.strip()
        + " | "
        + df["Secondary Functional Group"].fillna("").astype(str).str.strip()
        + " | "
        + df["Primary Specialization"].fillna("").astype(str).str.strip()
        + " | "
        + df["Secondary Specialization"].fillna("").astype(str).str.strip()
    )


def add_conditional_error_formatting(ws):
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import PatternFill
    from openpyxl.formatting.rule import FormulaRule

    max_row = ws.max_row
    max_col = ws.max_column
    if max_row < 2 or max_col < 1:
        return

    rng = f"A2:{get_column_letter(max_col)}{max_row}"
    formula = 'ISNUMBER(SEARCH("Error",A2))'
    rule = FormulaRule(formula=[formula], stopIfTrue=False,
                       fill=PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid"))
    ws.conditional_formatting.add(rng, rule)


def norm_error(msg: str) -> str:
    """Normalize error string for grouping (strip trailing details)."""
    return re.sub(r"->.*$", "", str(msg)).strip()


def compile_error_catalog(df: pd.DataFrame) -> pd.DataFrame:
    """Scan a dataframe for 'Error:' cells and summarize counts by message."""
    errs = []
    for c in df.columns:
        s = df[c].dropna().astype(str)
        errs.extend([x for x in s if x.startswith("Error:")])
    cnt = Counter(norm_error(e) for e in errs)
    if not cnt:
        return pd.DataFrame(columns=["Error_Message", "Count"])
    return pd.DataFrame([{"Error_Message": k, "Count": v} for k, v in cnt.most_common()])


def extract_unmapped_values(errors: list) -> dict:
    """
    From the errors list (with tails '-> value'), extract unmapped lookups.
    Returns dict of lists keyed by a label.
    """
    buckets = defaultdict(set)
    for e in errors:
        m = re.search(r"Error: (.+?) from (.+?) not found -> '(.+)'", e)
        if m:
            field, tab, val = m.groups()
            key = f"Unmapped_{field}_from_{tab}"
            buckets[key].add(val)
            continue
        m3 = re.search(r"Handler tab not found: ([A-Za-z0-9_]+)", e)
        if m3:
            buckets["Missing_Handler_Tabs"].add(m3.group(1))
            continue
        m4 = re.search(r"Role_Nm not found for key in ([A-Za-z0-9_]+): (.+)$", e)
        if m4:
            tab, key = m4.groups()
            buckets[f"Missing_Role_Key_in_{tab}"].add(key)
            continue
    return {k: sorted(v) for k, v in buckets.items()}


def drop_truly_empty_rows(df: pd.DataFrame) -> pd.DataFrame:
    """Drop rows where all CORE_COLS are empty/NaN."""
    work = df.copy()
    for c in CORE_COLS:
        if c not in work.columns:
            work[c] = ""
    mask_all_empty = work[CORE_COLS].apply(lambda r: all(str(x).strip() == "" for x in r.values), axis=1)
    return df[~mask_all_empty].copy()


def df_with_totals_by_pfg(master_df: pd.DataFrame, descending: bool = True) -> pd.DataFrame:
    """Return 'Count by PFG' table with a Grand Total row, sorted by Count."""
    if "Primary Functional Group" not in master_df.columns:
        return pd.DataFrame(columns=["Primary Functional Group","Count"])
    grp = (master_df
           .groupby("Primary Functional Group", dropna=False)
           .size()
           .reset_index(name="Count")
           .sort_values("Count", ascending=not descending, kind="mergesort"))
    total = pd.DataFrame([{"Primary Functional Group":"Grand Total", "Count": int(grp["Count"].sum())}])
    return pd.concat([grp, total], ignore_index=True)


def df_with_totals_by_pfg_role(master_df: pd.DataFrame, descending: bool = True) -> pd.DataFrame:
    """
    Return a block-style table: for each PFG, rows of ROLE_Nm sorted by Count (desc by default)
    and a 'Total' line per PFG.
    """
    needed_cols = ["Primary Functional Group","ROLE_Nm"]
    for c in needed_cols:
        if c not in master_df.columns:
            master_df[c] = ""
    base = (master_df
            .groupby(needed_cols, dropna=False)
            .size()
            .reset_index(name="Count"))
    # Build blocks sorted by Count within each PFG
    out_rows = []
    # Use mergesort to keep stability if equal counts; break ties by ROLE_Nm
    for pfg, sub in base.groupby("Primary Functional Group", dropna=False, sort=False):
        sub_sorted = sub.sort_values(
            by=["Count","ROLE_Nm"],
            ascending=[not descending, True],
            kind="mergesort"
        )
        for _, r in sub_sorted.iterrows():
            out_rows.append({
                "Primary Functional Group": r["Primary Functional Group"],
                "ROLE_Nm": r["ROLE_Nm"],
                "Count": int(r["Count"]),
            })
        out_rows.append({
            "Primary Functional Group": str(pfg) if pd.notna(pfg) else "",
            "ROLE_Nm": "Total",
            "Count": int(sub["Count"].sum()),
        })
    return pd.DataFrame(out_rows, columns=["Primary Functional Group","ROLE_Nm","Count"])


# =============================
# Main
# =============================

def main():
    parser = argparse.ArgumentParser(description="Roles Dyna - Phase 1 (v1.4_2019a, sorted summary, no-operator)")
    parser.add_argument("--config", default="config.json")
    parser.add_argument("--raw", default=None)
    parser.add_argument("--vars", default=None)
    parser.add_argument("--fgp", default="")
    parser.add_argument("--errors-only", action="store_true")
    parser.add_argument("--version", default="v_1_4_2019a")
    parser.add_argument("--outputs-dir", default=None)
    parser.add_argument("--logs-dir", default=None)
    parser.add_argument("--summary-sort", choices=["asc","desc"], default="desc",
                        help="Sort order for Summary counts (default: desc)")
    args = parser.parse_args()

    # Sensible relative defaults (can be overridden by config or CLI)
    base_dir = os.path.dirname(os.path.abspath(__file__))
    default_raw = os.path.abspath(os.path.join(base_dir, "../../Input/Raw_input_07012025.xlsx"))
    default_vars = os.path.abspath(os.path.join(base_dir, "../../Input/Var_Trns_Mstr.xlsx"))
    default_out = os.path.abspath(os.path.join(base_dir, "../../Output"))
    default_logs = os.path.abspath(os.path.join(base_dir, "../../Logs"))

    cfg = {}
    if os.path.exists(args.config):
        try:
            cfg = load_config(args.config)
        except Exception:
            cfg = {}

    tz_name = (cfg.get("timezone") if isinstance(cfg, dict) else None) or "America/New_York"
    now, now_tag = tz_now_str(tz_name)

    raw_path = args.raw or (cfg.get("paths", {}).get("raw_input") if isinstance(cfg, dict) else None) or default_raw
    vars_path = args.vars or (cfg.get("paths", {}).get("translation_master") if isinstance(cfg, dict) else None) or default_vars
    outputs_dir = args.outputs_dir or (cfg.get("paths", {}).get("outputs_dir") if isinstance(cfg, dict) else None) or default_out
    logs_dir = args.logs_dir or (cfg.get("paths", {}).get("logs_dir") if isinstance(cfg, dict) else None) or default_logs
    fgp_filter = args.fgp or (cfg.get("options", {}).get("fgp_filter") if isinstance(cfg, dict) else "") or ""
    errors_only = bool(args.errors_only or ((cfg.get("options", {}) if isinstance(cfg, dict) else {}).get("errors_only", False)))
    version = args.version
    descending = (args.summary_sort == "desc")

    os.makedirs(outputs_dir, exist_ok=True)
    os.makedirs(logs_dir, exist_ok=True)

    run_log_lines = []
    def log(line: str):
        ts = now.strftime("%Y-%m-%d %H:%M:%S")
        run_log_lines.append(f"[{ts}] {line}")

    log("Run start")
    log(f"Version: {version}")
    log(f"Raw: {raw_path}")
    log(f"Vars: {vars_path}")
    log(f"FGP filter: '{fgp_filter}'")
    log(f"Summary sort: {'descending' if descending else 'ascending'}")

    # --- Load translation master ---
    var_sheets = read_all_sheets(vars_path)
    Title_Normalization = var_sheets.get("Title_Normalization")
    PFGC_Conv = var_sheets.get("PFGC_Conv")
    PSP_Conv = var_sheets.get("PSP_Conv")
    SSPL_Conv = var_sheets.get("SSPL_Conv")
    Loc_Sz = var_sheets.get("Loc_Sz")
    ADDL_Keys = var_sheets.get("ADDL_Keys")

    # Maps
    pfg_map = {}
    if PFGC_Conv is not None:
        PFGC_Conv = normalize_cols(PFGC_Conv)
        cols = {c.lower(): c for c in PFGC_Conv.columns}
        long_col = cols.get("orig_primary_functional_group") or cols.get("primary functional group") or list(PFGC_Conv.columns)[0]
        short_col = cols.get("shrt_func") or list(PFGC_Conv.columns)[1]
        pfg_map = dict(zip(PFGC_Conv[long_col].fillna("").astype(str).str.strip(),
                           PFGC_Conv[short_col].fillna("").astype(str).str.strip()))

    psp_map = {}
    if PSP_Conv is not None:
        PSP_Conv = normalize_cols(PSP_Conv)
        cols = {c.lower(): c for c in PSP_Conv.columns}
        long_col = cols.get("primary specialization") or list(PSP_Conv.columns)[0]
        short_col = cols.get("shrtpspec") or list(PSP_Conv.columns)[1]
        psp_map = dict(zip(PSP_Conv[long_col].fillna("").astype(str).str.strip(),
                           PSP_Conv[short_col].fillna("").astype(str).str.strip()))

    sspl_map = {}
    if SSPL_Conv is not None:
        SSPL_Conv = normalize_cols(SSPL_Conv)
        cols = {c.lower(): c for c in SSPL_Conv.columns}
        long_col = cols.get("secondary specialization") or list(SSPL_Conv.columns)[0]
        short_col = cols.get("shrt_sspl") or list(SSPL_Conv.columns)[1]
        sspl_map = dict(zip(SSPL_Conv[long_col].fillna("").astype(str).str.strip(),
                            SSPL_Conv[short_col].fillna("").astype(str).str.strip()))

    title_map = make_title_normalizer(Title_Normalization) if Title_Normalization is not None else {}

    # Location size map
    loc_size_map = {}
    if Loc_Sz is not None:
        df_ls = normalize_cols(Loc_Sz)
        cols = {c.lower(): c for c in df_ls.columns}
        loc_col = cols.get("location") or list(df_ls.columns)[0]
        size_col = cols.get("size") or list(df_ls.columns)[1]
        loc_size_map = dict(
            zip(
                df_ls[loc_col].fillna("").astype(str).str.strip(),
                df_ls[size_col].fillna("").astype(str).str.strip(),
            )
        )

    # ADDL Keys
    addl_keys = []
    if ADDL_Keys is not None:
        df_k = normalize_cols(ADDL_Keys)
        if "Key_name" in df_k.columns:
            addl_keys = [k for k in df_k["Key_name"].dropna().astype(str).str.strip().tolist() if k]

    # --- RAW input: stitch tables from all sheets ---
    raw_book = read_all_sheets(raw_path)
    tables = []
    rejected_all = []
    for sh_name, df in raw_book.items():
        df = df.fillna("")
        header_idx, cols = detect_header_row(df)
        if header_idx is None:
            log(f"[{sh_name}] No header detected; skipping sheet.")
            continue
        tbl = slice_table_from_header(df, header_idx, cols)
        tbl = coalesce_title_column(tbl)
        tbl = drop_repeated_headers(tbl)
        tbl, rejected = filter_reject_rows(tbl)
        tbl["__source_sheet"] = sh_name
        rejected["__source_sheet"] = sh_name
        tables.append(tbl)
        if not rejected.empty:
            rejected_all.append(rejected)

    if not tables:
        raise SystemExit("No valid tables detected in RAW input.")

    raw_df = pd.concat(tables, ignore_index=True)
    raw_df = drop_truly_empty_rows(raw_df)

    input_rows = len(raw_df)

    # Optional filter by FGP (accept long or short form)
    if fgp_filter:
        fgp_lower = fgp_filter.strip().lower()
        raw_df = raw_df[
            raw_df.get("Primary Functional Group","").astype(str).str.lower().eq(fgp_lower)
            | raw_df.get("Primary Functional Group","").map(lambda x: pfg_map.get(str(x).strip(), "")).str.lower().eq(fgp_lower)
        ]

    # Title normalization
    errors = []
    if title_map and "Position/Title" in raw_df.columns:
        raw_df["Position/Title"] = raw_df["Position/Title"].apply(lambda x: title_map.get(str(x).strip(), str(x).strip()))

    # Short forms
    if "Primary Functional Group" in raw_df.columns and pfg_map:
        raw_df["Shrt_Func"] = map_with_errors(
            raw_df["Primary Functional Group"], pfg_map, "Primary Functional Group", "PFGC_Conv", errors
        )
    else:
        raw_df["Shrt_Func"] = ""

    if "Primary Specialization" in raw_df.columns and psp_map:
        raw_df["ShrtPSpec"] = map_with_errors(
            raw_df["Primary Specialization"], psp_map, "Primary Specialization", "PSP_Conv", errors
        )
    else:
        raw_df["ShrtPSpec"] = ""

    if "Secondary Specialization" in raw_df.columns and sspl_map:
        raw_df["Shrt_SSpl"] = map_with_errors(
            raw_df["Secondary Specialization"], sspl_map, "Secondary Specialization", "SSPL_Conv", errors
        )
    else:
        raw_df["Shrt_SSpl"] = ""

    # Location size
    if loc_size_map and "Location" in raw_df.columns:
        raw_df["Location_Size"] = raw_df["Location"].apply(lambda x: loc_size_map.get(str(x).strip(), ""))

    # Build handler key
    raw_df["__hndl_key"] = build_hndl_key(raw_df)

    # =============================
    # Vectorized Role Lookup (cached handler tabs) WITH DEDUP
    # =============================
    raw_df["__pf_short"] = raw_df.apply(
        lambda r: pfg_map.get(str(r.get("Primary Functional Group","")).strip(), str(r.get("Shrt_Func","")).strip()), axis=1
    )

    handlers_all = []
    missing_tabs = set()
    per_tab_conflicts = {}

    for pf_short in sorted(set([x for x in raw_df["__pf_short"].astype(str).tolist() if x])):
        tab_name = f"{pf_short}_Hndl"
        hndl_df = var_sheets.get(tab_name)
        if hndl_df is None:
            missing_tabs.add(tab_name)
            continue
        hndl_df = normalize_cols(hndl_df)
        hndl_df = coalesce_title_column(hndl_df)
        hndl_df["__hndl_key"] = build_hndl_key(hndl_df)

        role_col_name = None
        for c in hndl_df.columns:
            lc = c.strip().lower()
            if lc in ("role_nm","role name","role"):
                role_col_name = c
                break
        if role_col_name is None:
            hndl_df["ROLE_Nm"] = ""
        else:
            hndl_df.rename(columns={role_col_name: "ROLE_Nm"}, inplace=True)

        # Detect duplicates per tab
        dup = hndl_df.duplicated(subset="__hndl_key", keep=False)
        if dup.any():
            per_tab_conflicts[tab_name] = (
                hndl_df.loc[dup, ["__hndl_key","ROLE_Nm"]]
                .groupby("__hndl_key")
                .agg(unique_roles=("ROLE_Nm", lambda s: sorted(set([str(x).strip() for x in s if str(x).strip()!='']))),
                     count=("ROLE_Nm","size"))
                .reset_index()
            )

        handlers_all.append(hndl_df[["__hndl_key", "ROLE_Nm"]])

    if handlers_all:
        handlers_all_df = pd.concat(handlers_all, ignore_index=True)

        # GLOBAL dedup: prefer non-empty ROLE_Nm
        tmp = handlers_all_df.copy()
        tmp["__is_empty"] = tmp["ROLE_Nm"].fillna("").astype(str).str.strip().eq("")
        tmp.sort_values(["__hndl_key","__is_empty","ROLE_Nm"], inplace=True)
        handlers_all_unique = tmp.drop_duplicates("__hndl_key", keep="first").drop(columns="__is_empty")
    else:
        handlers_all_unique = pd.DataFrame(columns=["__hndl_key","ROLE_Nm"])

    # Merge once (no row explosion now)
    merged = raw_df.merge(handlers_all_unique, on="__hndl_key", how="left", suffixes=("", "_hndl"))
    merged_rows = len(merged)

    # Fill ROLE_Nm with appropriate errors
    def role_error_for_row(row):
        pf_short = str(row.get("__pf_short","")).strip()
        if not pf_short:
            return "Error: Primary Functional Group short code missing"
        tab_name = f"{pf_short}_Hndl"
        if tab_name in missing_tabs:
            return f"Error: Handler tab not found: {tab_name}"
        if pd.isna(row.get("ROLE_Nm","")) or str(row.get("ROLE_Nm","")).strip() == "":
            return f"Error: Role_Nm not found in {tab_name}"
        return str(row["ROLE_Nm"]).strip()

    merged["ROLE_Nm"] = merged.apply(role_error_for_row, axis=1)

    # ADDL Keys placeholders (create columns but leave blank in Phase 1)
    for k in (ADDL_Keys["Key_name"].dropna().astype(str).str.strip().tolist() if isinstance(ADDL_Keys, pd.DataFrame) and "Key_name" in ADDL_Keys.columns else []):
        if k not in merged.columns:
            merged[k] = ""

    # Final master
    master_df = merged.drop(columns=["__pf_short"], errors="ignore").copy()
    master_df = drop_truly_empty_rows(master_df)

    # Exceptions detection
    excp_mask = master_df.apply(lambda r: any(isinstance(v, str) and v.startswith("Error:") for v in r.values), axis=1)
    excp_rows = master_df[excp_mask].copy()

    if errors_only:
        master_df = master_df[excp_mask].copy()

    # ----- Build Summary tables for Excel 2019 (sorted) -----
    sum_by_pfg = df_with_totals_by_pfg(master_df, descending=descending)
    sum_by_pfg_role = df_with_totals_by_pfg_role(master_df, descending=descending)

    # Lists for validation pickers
    pfg_list = sorted([x for x in master_df.get("Primary Functional Group", pd.Series(dtype=str)).fillna("").astype(str).unique() if x])
    role_list = sorted([x for x in master_df.get("ROLE_Nm", pd.Series(dtype=str)).fillna("").astype(str).unique() if x])

    # Run_Rep
    run_rep = pd.DataFrame(
        [{
            "Run_Timestamp": now.strftime("%m%d%Y_%H_%M"),
            "Timezone": tz_name,
            "Version": version,
            "Raw_File": raw_path,
            "Var_Master_File": vars_path,
            "FGP_Filter": fgp_filter,
            "Input_Rows": input_rows,
            "Master_Records_Written": len(master_df),
            "Merged_Rows_Before_Filter": merged_rows,
            "Exception_Rows": len(excp_rows),
            "Summary_Sort": "desc" if descending else "asc",
        }]
    )

    # Error Catalog + Unmapped tabs
    error_catalog = compile_error_catalog(master_df)

    mapping_errors = []
    for c in master_df.columns:
        s = master_df[c].dropna().astype(str)
        mapping_errors.extend([x for x in s if x.startswith("Error:")])

    unmapped_dict = extract_unmapped_values(mapping_errors)
    unmapped_tabs = {k: pd.DataFrame({k: v}) for k, v in unmapped_dict.items()}

    # Write Excel
    out_name = f"Roles_Output_{version}_{now.strftime('%d%m%y_%H_%M')}.xlsx"
    out_path = os.path.join(outputs_dir, out_name)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        # Master & friends
        master_df.to_excel(writer, index=False, sheet_name="Master_Roles_by_User")
        run_rep.to_excel(writer, index=False, sheet_name="Run_Rep")
        excp_rows.to_excel(writer, index=False, sheet_name="Excp_Rep")
        error_catalog.to_excel(writer, index=False, sheet_name="Error_Catalog")
        for tab, df_tab in unmapped_tabs.items():
            sheet_name = tab[:31]
            df_tab.to_excel(writer, index=False, sheet_name=sheet_name)

        # Summary layout
        sum_sheet = writer.book.create_sheet("Summary")
        writer.sheets["Summary"] = sum_sheet

        # Controls labels
        sum_sheet["A1"] = "Summary (Excel 2019 friendly)"
        sum_sheet["A2"] = "Primary Functional Group (picker)"
        sum_sheet["A3"] = "ROLE_Nm (picker)"
        sum_sheet["A5"] = "Selected users total (filters on Records)"
        sum_sheet["A7"] = "Open filtered records:"
        # We'll set B5 to SUBTOTAL later after creating the Records table
        sum_sheet["B7"] = "Click to see records"
        sum_sheet["B7"].hyperlink = "#Records!A1"
        try:
            from openpyxl.styles import Font
            sum_sheet["B7"].font = Font(color="0000EE", underline="single")
        except Exception:
            pass

        # Place the two summary tables
        # Table 1: Count by PFG (with Grand Total)
        start_row_1 = 10
        for j, col in enumerate(["Primary Functional Group","Count"], start=1):
            sum_sheet.cell(row=start_row_1, column=j, value=col)
        for i, row in sum_by_pfg.iterrows():
            sum_sheet.cell(row=start_row_1 + 1 + i, column=1, value=row["Primary Functional Group"])
            sum_sheet.cell(row=start_row_1 + 1 + i, column=2, value=int(row["Count"]) if pd.notna(row["Count"]) else None)

        # Table 2: Count by PFG + Role (with Totals per PFG)
        start_row_2 = start_row_1 + len(sum_by_pfg) + 4
        for j, col in enumerate(["Primary Functional Group","ROLE_Nm","Count"], start=1):
            sum_sheet.cell(row=start_row_2, column=j, value=col)
        for i, row in sum_by_pfg_role.iterrows():
            sum_sheet.cell(row=start_row_2 + 1 + i, column=1, value=row["Primary Functional Group"])
            sum_sheet.cell(row=start_row_2 + 1 + i, column=2, value=row["ROLE_Nm"])
            sum_sheet.cell(row=start_row_2 + 1 + i, column=3, value=int(row["Count"]) if pd.notna(row["Count"]) else None)

        # Records sheet (as Excel Table with AutoFilter)
        # Use pandas to write, then convert to Table
        master_df.to_excel(writer, index=False, sheet_name="Records")
        ws_rec = writer.sheets["Records"]

        # Create Excel Table
        from openpyxl.worksheet.table import Table, TableStyleInfo
        max_row = ws_rec.max_row
        max_col = ws_rec.max_column
        last_col_letter = ws_rec.cell(row=1, column=max_col).column_letter
        table_ref = f"A1:{last_col_letter}{max_row}"
        table = Table(displayName="T_Records", ref=table_ref)
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws_rec.add_table(table)

        # Selected users total (SUBTOTAL counts visible names after filter)
        sum_sheet["B5"] = "=SUBTOTAL(103, T_Records[Name])"

        # Lists sheet for pickers
        ws_lists = writer.book.create_sheet("Lists")
        writer.sheets["Lists"] = ws_lists
        ws_lists["A1"] = "PFG_List"
        for i, v in enumerate(pfg_list, start=2):
            ws_lists.cell(row=i, column=1, value=v)
        ws_lists["B1"] = "ROLE_List"
        for i, v in enumerate(role_list, start=2):
            ws_lists.cell(row=i, column=2, value=v)

        # Data validation for pickers (2019 compatible)
        from openpyxl.worksheet.datavalidation import DataValidation
        dv_pfg = DataValidation(type="list", formula1="=Lists!$A$2:$A${}".format(1 + len(pfg_list)), allow_blank=True)
        dv_role = DataValidation(type="list", formula1="=Lists!$B$2:$B${}".format(1 + len(role_list)), allow_blank=True)
        sum_sheet.add_data_validation(dv_pfg)
        sum_sheet.add_data_validation(dv_role)
        dv_pfg.add(sum_sheet["B2"])
        dv_role.add(sum_sheet["B3"])

        # Stash raw data on Summary!AA1
        # AA is col 27
        start_col_raw = 27
        # headers
        for j, col in enumerate(master_df.columns.tolist(), start=start_col_raw):
            sum_sheet.cell(row=1, column=j, value=col)
        # rows
        for i, (_, r) in enumerate(master_df.iterrows(), start=2):
            for j, col in enumerate(master_df.columns.tolist(), start=start_col_raw):
                sum_sheet.cell(row=i, column=j, value=r[col])

        # Conditional formatting on Master
        ws_master = writer.sheets["Master_Roles_by_User"]
        add_conditional_error_formatting(ws_master)

        # Optional: hide Lists sheet
        try:
            ws_lists.sheet_state = "hidden"
        except Exception:
            pass

    # Append to external log (append at bottom for Phase 1)
    log_path = os.path.join(logs_dir, "Roles_Dyn_cons_Log.txt")
    try:
        mode = "a" if os.path.exists(log_path) else "w"
        with open(log_path, mode, encoding="utf-8") as lf:
            lf.write("\n".join(run_log_lines) + "\n")
    except Exception as e:
        print(f"WARNING: could not write log file: {e}")

    print(f"Output written: {out_path}")
    print(f"Input rows: {input_rows}  |  After merge: {merged_rows}  |  Master written: {len(master_df)}")
    print(f"Exceptions: {len(excp_rows)}")


if __name__ == "__main__":
    main()
